# api/convert.py
import sys
import logging

# Add logging to see what's available
logging.basicConfig(level=logging.INFO)
logging.info(f"Python version: {sys.version}")

# Try importing the libraries
try:
    import openpyxl
    logging.info(f"openpyxl version: {openpyxl.__version__}")
except ImportError as e:
    logging.error(f"Failed to import openpyxl: {e}")

try:
    import reportlab
    logging.info(f"reportlab version: {reportlab.__version__}")
except ImportError as e:
    logging.error(f"Failed to import reportlab: {e}")
    
from http.server import BaseHTTPRequestHandler
import json
import io
import base64
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            # Get content length
            content_length = int(self.headers['Content-Length'])
            
            # Read the Excel file from request body
            excel_data = self.rfile.read(content_length)
            
            # Load Excel workbook
            wb = load_workbook(io.BytesIO(excel_data), data_only=True)
            sheet = wb.active
            
            # Create PDF in memory
            pdf_buffer = io.BytesIO()
            doc = SimpleDocTemplate(pdf_buffer, pagesize=letter)
            elements = []
            
            # Extract data from Excel
            data = []
            for row in sheet.iter_rows(values_only=True):
                clean_row = [str(cell) if cell is not None else '' for cell in row]
                data.append(clean_row)
            
            if data:
                # Create table
                table = Table(data)
                
                # Style the table
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            # Get PDF bytes and encode to base64
            pdf_bytes = pdf_buffer.getvalue()
            pdf_buffer.close()
            pdf_base64 = base64.b64encode(pdf_bytes).decode('utf-8')
            
            # Send response
            self.send_response(200)
            self.send_header('Content-type', 'text/plain')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            self.wfile.write(pdf_base64.encode())
            
        except Exception as e:
            self.send_response(500)
            self.send_header('Content-type', 'application/json')
            self.end_headers()
            error_response = json.dumps({'error': str(e)})
            self.wfile.write(error_response.encode())
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
